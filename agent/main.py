import pandas as pd
import json
import random
import sys
import tempfile
import os
from typing import Dict, Any, List, Tuple, Optional
from openpyxl import load_workbook
from io import BytesIO
from datetime import datetime

# Google Cloud imports
from google.cloud import firestore
from google.cloud import storage
from google.oauth2 import service_account

# Try to import Google GenAI library
try:
    from google import genai
    from google.genai import types
    GENAI_AVAILABLE = True
except ImportError:
    GENAI_AVAILABLE = False
    print("Warning: google-genai library not found. Install with: pip install google-genai")

try:
    import PyPDF2
except ImportError:
    PyPDF2 = None
    print("Warning: PyPDF2 not found. Install with: pip install PyPDF2")

from dotenv import load_dotenv
load_dotenv()

# Environment variables
PROJECT_ID = os.environ.get("PROJECT_ID")
STORAGE_BUCKET = os.environ.get("STORAGE_BUCKET")
FIREBASE_SERVICE_ACCOUNT = os.environ.get("FIREBASE_SERVICE_ACCOUNT")
GOOGLE_API_KEY = os.environ.get("GOOGLE_API_KEY")


def get_firestore_client():
    """Return Firestore client using service account JSON from env."""
    if not FIREBASE_SERVICE_ACCOUNT or not PROJECT_ID:
        raise RuntimeError("FIREBASE_SERVICE_ACCOUNT or PROJECT_ID env var not set")

    credentials = service_account.Credentials.from_service_account_info(
        json.loads(FIREBASE_SERVICE_ACCOUNT)
    )
    return firestore.Client(credentials=credentials, project=PROJECT_ID, database='ats-db')


def get_storage_client():
    """Return Firebase Storage client using service account JSON from env."""
    if not FIREBASE_SERVICE_ACCOUNT or not STORAGE_BUCKET:
        raise RuntimeError("FIREBASE_SERVICE_ACCOUNT or STORAGE_BUCKET env var not set")
    
    credentials = service_account.Credentials.from_service_account_info(
        json.loads(FIREBASE_SERVICE_ACCOUNT)
    )
    return storage.Client(credentials=credentials, project=PROJECT_ID)


def excel_to_json(excel_path: str, output_json_path: str = None) -> Dict[str, Dict[str, Any]]:
    """
    Convert Excel file to JSON format.
    
    Removes empty rows and columns before conversion.
    Assumes:
    - First column contains row names/identifiers
    - First row contains column headers
    
    Args:
        excel_path: Path to the Excel file
        output_json_path: Optional path to save JSON file. If None, returns data only.
    
    Returns:
        Dictionary where keys are row names and values are dictionaries with column names as keys
        Format: {row_name: {col_name: value, ...}, ...}
    """
    # Read Excel file
    df = pd.read_excel(excel_path, header=None)
    
    # Remove completely empty rows
    df = df.dropna(how='all')
    
    # Remove completely empty columns
    df = df.dropna(axis=1, how='all')
    
    # Reset index after dropping rows
    df = df.reset_index(drop=True)
    
    # First row should be column headers
    if len(df) == 0:
        return {}
    
    # Set first row as column headers
    df.columns = df.iloc[0]
    df = df[1:].reset_index(drop=True)
    
    # First column should be row names
    if len(df.columns) == 0:
        return {}
    
    # Set first column as index (row names)
    row_name_col = df.columns[0]
    df = df.set_index(row_name_col)
    
    # Remove any rows where the row name is empty/NaN
    df = df[df.index.notna()]
    
    # Convert to nested dictionary structure
    result = {}
    for row_name, row_data in df.iterrows():
        row_dict = {}
        
        # Add all column values
        for col_name, value in row_data.items():
            # Convert NaN to None for JSON serialization
            if pd.isna(value):
                row_dict[str(col_name)] = None
            else:
                row_dict[str(col_name)] = value
        
        # Use row name as key in result dictionary
        result[str(row_name) if pd.notna(row_name) else ""] = row_dict
    
    # Save to JSON if output path provided
    if output_json_path:
        with open(output_json_path, 'w', encoding='utf-8') as f:
            json.dump(result, f, indent=2, ensure_ascii=False, default=str)
    
    return result


def json_to_excel_template(json_data: Dict[str, Any], template_path: str, output_excel_path: str) -> None:
    """
    Fill JSON data into a template Excel file, preserving the exact template structure.
    
    This function:
    1. Reads the template Excel file to understand its structure (including empty rows/columns)
    2. Uses the JSON data directly (not from file)
    3. Maps JSON data back into the template structure, preserving all empty rows and columns
    
    Args:
        json_data: Dictionary with the data to fill
        template_path: Path to the template Excel file
        output_excel_path: Path to save the filled Excel file
    """
    if not json_data or not isinstance(json_data, dict):
        # If no JSON data, just copy the template
        import shutil
        shutil.copy(template_path, output_excel_path)
        return
    
    # Load template workbook
    wb = load_workbook(template_path)
    ws = wb.active
    
    # Find the header row (first row with column names)
    # Look for the first row that has non-empty cells
    header_row_idx = None
    max_col = ws.max_column
    
    for row_idx in range(1, ws.max_row + 1):
        row_values = [ws.cell(row=row_idx, column=col).value for col in range(1, max_col + 1)]
        if any(v is not None and str(v).strip() != '' for v in row_values):
            header_row_idx = row_idx
            break
    
    if header_row_idx is None:
        # No header found, just copy template
        wb.save(output_excel_path)
        return
    
    # Find row name column first
    row_name_col = 2  # Default to column B
    
    # Check if column 1 has values, if not use column 2
    has_col1_values = any(
        ws.cell(row=row_idx, column=1).value is not None 
        for row_idx in range(header_row_idx + 1, min(header_row_idx + 10, ws.max_row + 1))
    )
    if has_col1_values:
        row_name_col = 1
    
    # Extract column headers from header row
    # Start from the column after the row name column
    column_headers = {}  # col_idx -> original_value
    start_col = row_name_col + 1
    for col_idx in range(start_col, max_col + 1):
        cell_value = ws.cell(row=header_row_idx, column=col_idx).value
        if cell_value is not None and str(cell_value).strip() != '':
            column_headers[col_idx] = cell_value
    
    # Helper function to normalize column names for matching
    def normalize_col_name(value):
        """Convert column value to string, handling numeric values."""
        if value is None:
            return None
        try:
            # If it's a number, try to match both with and without .0
            num_val = float(value)
            return str(num_val)
        except:
            return str(value).strip()
    
    # Get sample JSON row to understand column name format
    sample_json_row = next(iter(json_data.values())) if json_data else {}
    json_col_names = set(sample_json_row.keys()) if isinstance(sample_json_row, dict) else set()
    
    # Create mapping: JSON column name -> Excel column index
    json_to_col = {}
    for col_idx, col_value in column_headers.items():
        normalized_template = normalize_col_name(col_value)
        
        # Try to find matching JSON column name
        for json_col_name in json_col_names:
            normalized_json = normalize_col_name(json_col_name)
            
            # Try exact match
            if normalized_template == normalized_json:
                json_to_col[json_col_name] = col_idx
                break
            # Try with .0 suffix variations
            try:
                if normalized_template.replace('.0', '') == normalized_json.replace('.0', ''):
                    json_to_col[json_col_name] = col_idx
                    break
            except:
                pass
    
    # Find row names and map them
    # Row names are in the identified row_name_col, starting from row after header
    row_name_to_row_idx = {}
    for row_idx in range(header_row_idx + 1, ws.max_row + 1):
        cell_value = ws.cell(row=row_idx, column=row_name_col).value
        if cell_value is not None and str(cell_value).strip() != '':
            row_name = str(cell_value).strip()
            row_name_to_row_idx[row_name] = row_idx
    
    # Fill in the data from JSON
    for json_row_name, json_row_data in json_data.items():
        if not isinstance(json_row_data, dict):
            continue
        
        # Find matching row in template
        row_idx = None
        # Try exact match first
        if json_row_name in row_name_to_row_idx:
            row_idx = row_name_to_row_idx[json_row_name]
        else:
            # Try case-insensitive match
            for template_row_name, template_row_idx in row_name_to_row_idx.items():
                if json_row_name.lower() == template_row_name.lower():
                    row_idx = template_row_idx
                    break
        
        if row_idx is None:
            # Row not found in template, skip it
            continue
        
        # Fill in column values
        for json_col_name, json_value in json_row_data.items():
            # Find matching column using the mapping
            if json_col_name in json_to_col:
                col_idx = json_to_col[json_col_name]
                # Write the value to the cell
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = json_value
    
    # Save the filled workbook
    wb.save(output_excel_path)


def create_simplified_schema(template_json: Dict[str, Any]) -> Dict[str, Any]:
    """
    Creates a simplified JSON schema that defines the structure of a row,
    but doesn't hardcode every single row name as a property key.
    
    Args:
        template_json: The template dictionary to extract column structure from.
        
    Returns:
        A JSON schema dictionary for a list of row objects.
    """
    # Extract unique column names (assuming uniform structure across most rows)
    # Find the first row that is a dictionary to get column names
    col_properties = {}
    col_required = []
    
    found_structure = False
    for row_data in template_json.values():
        if isinstance(row_data, dict):
            for col_name in row_data.keys():
                col_required.append(col_name)
                # Use NUMBER for financial values, allow null
                col_properties[col_name] = {
                    "type": "NUMBER",
                    "nullable": True
                }
            found_structure = True
            break
    
    if not found_structure:
        # Fallback if no dict rows found
        return {"type": "OBJECT", "properties": {}, "nullable": True}

    # Define schema for a single row item
    row_item_schema = {
        "type": "OBJECT",
        "properties": {
            "row_name": {
                "type": "STRING",
                "description": "The exact name of the row as requested."
            },
            "values": {
                "type": "OBJECT",
                "properties": col_properties,
                "required": col_required,
                "nullable": True
            }
        },
        "required": ["row_name", "values"]
    }

    # Define the root schema as an object containing a list of rows
    schema = {
        "type": "OBJECT",
        "properties": {
            "financial_data": {
                "type": "ARRAY",
                "items": row_item_schema
            }
        },
        "required": ["financial_data"]
    }
    
    return schema


def generate_solution_from_template_and_pdfs(template_json: Dict[str, Dict[str, Any]], 
                                             pdf_data: Dict[str, Any] = None) -> Dict[str, Dict[str, Any]]:
    """
    Generate a solution JSON structure from template and parsed PDF data using Gemini API.
    
    Args:
        template_json: Template JSON structure with row names and column names
        pdf_data: Dictionary with extracted data from PDFs
    
    Returns:
        JSON structure with same keys but populated with solution values
    """
    if not GOOGLE_API_KEY:
        print("Warning: GOOGLE_API_KEY not found in environment. Falling back to random generation.")
        return _generate_random_placeholder(template_json)
        
    if not GENAI_AVAILABLE:
        print("Warning: google-genai library not available. Falling back to random generation.")
        return _generate_random_placeholder(template_json)

    if not pdf_data:
        print("Warning: No PDF data provided. Falling back to random generation.")
        return _generate_random_placeholder(template_json)

    print("Generating solution using Gemini API...")
    
    try:
        client = genai.Client(api_key=GOOGLE_API_KEY)
        
        # Prepare context from PDFs
        pdf_context = ""
        for filename, text in pdf_data.items():
            pdf_context += f"\n--- Start of {filename} ---\n{text}\n--- End of {filename} ---\n"
            
        # Construct simplified schema based on template columns
        json_schema = create_simplified_schema(template_json)
        
        # Extract row names to guide the LLM
        row_names_list = list(template_json.keys())
        row_names_str = "\n".join([f"- {name}" for name in row_names_list])
        
        prompt = f"""
        You are an expert financial analyst. Your task is to extract financial data from the provided documents and populate a structured JSON list.
        
        INSTRUCTIONS:
        1. Read the provided document text carefully.
        2. Extract values for each row listed below.
        3. For each row, populate the values for the defined years/columns.
        4. If a value is explicitly mentioned, use it.
        5. If a value can be calculated (e.g., Total = Sum of parts), calculate it.
        6. If a value is not found, use null.
        7. OUTPUT FORMAT: A JSON object with a key "financial_data" containing a list of row objects.
           Each row object must have "row_name" and "values".
        
        REQUIRED ROWS TO EXTRACT:
        {row_names_str}
        
        DOCUMENT TEXT:
        {pdf_context[:50000]}
        
        Generate the JSON response matching the schema.
        """
        
        # Call Gemini API
        response = client.models.generate_content(
            model="gemini-2.5-flash",
            contents=prompt,
            config={
                "response_mime_type": "application/json",
                "response_schema": json_schema,
            },
        )
        
        # Parse response and convert back to template structure
        if response.text:
            generated_data_raw = json.loads(response.text)
            
            # Convert list format back to dictionary format {row_name: {col: val}}
            generated_data = {}
            
            # Helper to quickly find row data
            rows_list = generated_data_raw.get("financial_data", [])
            row_map = {item.get("row_name"): item.get("values") for item in rows_list}
            
            # Fill the template structure
            for row_name in template_json.keys():
                if row_name in row_map:
                    generated_data[row_name] = row_map[row_name]
                else:
                    # Row missing in output, keep as None or empty dict matching template type
                    if isinstance(template_json[row_name], dict):
                         # Initialize with nulls for columns if missing
                         generated_data[row_name] = {k: None for k in template_json[row_name].keys()}
                    else:
                        generated_data[row_name] = None
                        
            print("Successfully generated data using LLM.")
            return generated_data
        else:
            print("Error: Empty response from LLM.")
            return _generate_random_placeholder(template_json)
            
    except Exception as e:
        print(f"Error calling Gemini API: {str(e)}")
        import traceback
        traceback.print_exc()
        print("Falling back to random generation.")
        return _generate_random_placeholder(template_json)


def _generate_random_placeholder(template_json: Dict[str, Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    """Fallback function for random generation if API fails."""
    generated_json = {}
    
    for row_name, row_data in template_json.items():
        if not isinstance(row_data, dict):
            generated_json[row_name] = row_data
            continue
        
        generated_row = {}
        for col_name, value in row_data.items():
            if value is None:
                generated_row[col_name] = round(random.uniform(0, 1000000), 2)
            else:
                generated_row[col_name] = value
        
        generated_json[row_name] = generated_row
    
    return generated_json


def parse_pdfs_from_bytes(pdf_bytes_list: List[Tuple[str, bytes]]) -> Dict[str, Any]:
    """
    Extract all text from PDF files provided as bytes.
    
    Args:
        pdf_bytes_list: List of tuples (filename, pdf_bytes)
    
    Returns:
        Dictionary with extracted text from PDFs, keyed by filename
        Format: {filename: extracted_text, ...}
    """
    if PyPDF2 is None:
        raise ImportError("PyPDF2 is required for PDF parsing. Install it with: pip install PyPDF2")
    
    pdf_data = {}
    
    for filename, pdf_bytes in pdf_bytes_list:
        try:
            # Create a BytesIO object from the bytes
            pdf_file = BytesIO(pdf_bytes)
            pdf_reader = PyPDF2.PdfReader(pdf_file)
            
            # Extract text from all pages
            extracted_text = ""
            for page_num in range(len(pdf_reader.pages)):
                page = pdf_reader.pages[page_num]
                page_text = page.extract_text()
                extracted_text += page_text
            
            # Store extracted text with filename as key
            pdf_data[filename] = extracted_text
            print(f"Extracted {len(extracted_text)} characters from {filename}")
            
        except Exception as e:
            print(f"Error extracting text from {filename}: {str(e)}")
            # Store error message instead of text
            pdf_data[filename] = f"Error: {str(e)}"
    
    return pdf_data


def process_request(request_id: str):
    """
    Main function to process an extraction request.
    
    Steps:
    1. Get request details from Firestore
    2. Download template Excel and PDFs from Storage
    3. Parse template to JSON
    4. Parse PDFs to extract text
    5. Generate solution JSON using LLM
    6. Convert solution JSON to Excel using template structure
    7. Upload solution.xlsx to Storage
    8. Update Firestore request status to 'completed'
    
    Args:
        request_id: The Firestore document ID for the request
    """
    print(f"[PROCESS] Starting processing for request: {request_id}")
    
    # Initialize clients
    firestore_client = get_firestore_client()
    storage_client = get_storage_client()
    bucket_name = STORAGE_BUCKET.replace('gs://', '').strip()
    bucket = storage_client.bucket(bucket_name)
    
    # Get request document
    request_ref = firestore_client.collection('extraction_requests').document(request_id)
    request_doc = request_ref.get()
    
    if not request_doc.exists:
        raise RuntimeError(f"Request {request_id} not found in Firestore")
    
    request_data = request_doc.to_dict()
    print(f"[PROCESS] Request found: {request_data.get('template_filename')}, {request_data.get('pdf_count')} PDFs")
    
    # Update status to processing
    request_ref.update({
        'status': 'processing',
        'updated_at': datetime.utcnow()
    })
    print(f"[PROCESS] Updated status to 'processing'")
    
    try:
        # Create temporary directory for processing
        with tempfile.TemporaryDirectory() as temp_dir:
            print(f"[PROCESS] Created temp directory: {temp_dir}")
            
            # Download template file
            template_blob_path = request_data.get('template_blob_path')
            if not template_blob_path:
                raise RuntimeError("Template blob path not found in request")
            
            template_blob = bucket.blob(template_blob_path)
            if not template_blob.exists():
                raise RuntimeError(f"Template file not found in storage: {template_blob_path}")
            
            # Determine file extension from blob path or use .xlsx as default
            template_ext = template_blob_path.split('.')[-1] if '.' in template_blob_path else 'xlsx'
            template_local_path = os.path.join(temp_dir, f'template.{template_ext}')
            template_blob.download_to_filename(template_local_path)
            print(f"[PROCESS] Downloaded template: {template_blob_path} -> {template_local_path}")
            
            # Download PDF files
            pdf_blob_paths = request_data.get('pdf_blob_paths', [])
            if not pdf_blob_paths:
                raise RuntimeError("PDF blob paths not found in request")
            
            pdf_bytes_list = []
            for pdf_blob_path in pdf_blob_paths:
                pdf_blob = bucket.blob(pdf_blob_path)
                if not pdf_blob.exists():
                    print(f"[PROCESS] Warning: PDF file not found in storage: {pdf_blob_path}, skipping")
                    continue
                
                pdf_bytes = pdf_blob.download_as_bytes()
                # Extract original filename from blob path (remove prefix like "pdf_1_")
                filename = os.path.basename(pdf_blob_path)
                # If filename starts with "pdf_", try to extract original name
                if filename.startswith('pdf_') and '_' in filename:
                    # Remove "pdf_X_" prefix
                    parts = filename.split('_', 2)
                    if len(parts) >= 3:
                        filename = parts[2]
                pdf_bytes_list.append((filename, pdf_bytes))
                print(f"[PROCESS] Downloaded PDF: {filename} ({len(pdf_bytes)} bytes)")
            
            # Step 1: Convert template Excel to JSON
            print(f"[PROCESS] Step 1: Converting template to JSON...")
            template_json = excel_to_json(template_local_path)
            print(f"[PROCESS] Template has {len(template_json)} rows")
            
            # Step 2: Parse PDFs
            print(f"[PROCESS] Step 2: Parsing PDFs...")
            pdf_data = parse_pdfs_from_bytes(pdf_bytes_list)
            print(f"[PROCESS] Parsed {len(pdf_data)} PDFs")
            
            # Step 3: Generate solution JSON
            print(f"[PROCESS] Step 3: Generating solution using LLM...")
            generated_json = generate_solution_from_template_and_pdfs(template_json, pdf_data)
            print(f"[PROCESS] Generated solution with {len(generated_json)} rows")
            
            # Step 4: Convert solution JSON to Excel
            print(f"[PROCESS] Step 4: Converting solution to Excel...")
            solution_local_path = os.path.join(temp_dir, 'solution.xlsx')
            json_to_excel_template(generated_json, template_local_path, solution_local_path)
            print(f"[PROCESS] Created solution Excel file")
            
            # Step 5: Upload solution.xlsx to Storage
            print(f"[PROCESS] Step 5: Uploading solution to Storage...")
            solution_blob_path = f"{request_id}/solution.xlsx"
            solution_blob = bucket.blob(solution_blob_path)
            solution_blob.upload_from_filename(solution_local_path, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            print(f"[PROCESS] Uploaded solution to: {solution_blob_path}")
            
            # Step 6: Update Firestore status to completed
            print(f"[PROCESS] Step 6: Updating Firestore status...")
            request_ref.update({
                'status': 'completed',
                'solution_blob_path': solution_blob_path,
                'updated_at': datetime.utcnow()
            })
            print(f"[PROCESS] Successfully completed processing for request: {request_id}")
            
    except Exception as e:
        print(f"[PROCESS] ERROR processing request {request_id}: {e}")
        import traceback
        traceback.print_exc()
        
        # Update status to failed
        try:
            request_ref.update({
                'status': 'failed',
                'error': str(e),
                'updated_at': datetime.utcnow()
            })
        except:
            pass
        
        raise


if __name__ == "__main__":
    # Get request_id from command line argument or environment variable
    if len(sys.argv) > 1:
        request_id = sys.argv[1]
    else:
        request_id = os.environ.get("REQUEST_ID")
    
    if not request_id:
        print("ERROR: REQUEST_ID must be provided as command line argument or environment variable")
        sys.exit(1)
    
    print(f"[MAIN] Starting Cloud Run job for request: {request_id}")
    print(f"[MAIN] PROJECT_ID: {PROJECT_ID}")
    print(f"[MAIN] STORAGE_BUCKET: {STORAGE_BUCKET}")
    
    try:
        process_request(request_id)
        print(f"[MAIN] Job completed successfully for request: {request_id}")
        sys.exit(0)
    except Exception as e:
        print(f"[MAIN] Job failed for request {request_id}: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
